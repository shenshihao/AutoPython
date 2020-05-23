from openpyxl import load_workbook
import os

'''
这种方式只能用于替换xlsx格式的某个字符串
针对xls后缀的excel不能采用

'''


# 找到特定sheet中，特定字符,并替换字符串
def find_string_in_sheet(sheet, SourceStr, DesStr):
    for column in sheet.iter_cols():
        for cell2 in column:
            if cell2.value is not None:
                info2 = cell2.value
                if info2 == SourceStr:
                    cell2.value = DesStr
                    print(cell2.value)


def change_item(table):
    wb = load_workbook(table)
    all_sheets = wb.sheetnames
    print(all_sheets)

    list = [['70390000', '75340000'], ['2019', '2020']]

    for it in list:
        print(it)
        SourceStr = it[0]
        DesStr = it[1]
        for i in range(len(all_sheets)):
            sheet = wb[all_sheets[i]]
            find_string_in_sheet(sheet, SourceStr, DesStr)

    wb.save(table)


# 该path为你本地的路径
path = 'C:/Users/user/Downloads/python/xlsx/'
for root, dirs, files in os.walk(path):
    for file in files:
        change_item(path + file)
        # print()
# 1、替换表格中的字串
# 需要改进增加for循环用来遍历
# change_item(table)
